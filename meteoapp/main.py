from pandas import isnull
from random import randint
from meteoapp import data
import openpyxl
import pandas as pd
import matplotlib.pyplot as plt
from windrose import WindroseAxes


def find_file(i, city):
    months_value = 'meteoapp/cities/' + city + '/' + data.months_dict[i]
    wb = openpyxl.reader.excel.load_workbook(filename=months_value, data_only=True)
    wb.active = 0
    wc = wb.active
    return wc


def get_columns(wc, list):
    columns = []
    for i in range(2):
        c = []
        columns.append(c)
    i = 0
    for column in columns:
        for wc_item in wc[list[i]]:
            column.append(wc_item.value)
        i += 1
    for i in range(2):
        columns[i].pop(0)
    return columns


# список даних обраного міста та місяця
def create_data_list(wc):
    rows = []
    for i in range(1400):
        r = []
        rows.append(r)
    i = 1
    for row in rows:
        for col_name in data.col_names:
            row.append(wc[col_name + str(i)].value)
        i += 1
    return rows


# мовна локалізація (переклад)
def lan_localisation(lan, lan1, i, city):
    months_value = 'meteoapp/cities/' + city + '/' + data.months_dict[i]
    wb = openpyxl.reader.excel.load_workbook(filename=months_value, data_only=True)
    wb.active = 0
    wc = wb.active
    wc['A1'] = data.column1_tr[lan1]
    for i in range(2, 1490):
        for j in range(9):
            if wc['D' + str(i)].value == data.wind_tr[lan][j]:
                wc['D' + str(i)] = data.wind_tr[lan1][j]
    wb.save(months_value)


# корекція даних
def data_correction(i, city, interp_method, interp_value):
    months_value = 'meteoapp/cities/' + city + '/' + data.months_dict[i]
    wb = openpyxl.reader.excel.load_workbook(filename=months_value, data_only=True)
    wb.active = 0
    wc = wb.active
    columns = data_interpolation(wc, interp_method, interp_value)  # інтерполяція даних

    # запис в бд
    for i in range(2, 1440):
        if isnull(wc['F' + str(i)].value):  # заповнення порожніх полей колонки ww позначенням CL - тобто, "без явищ"
            wc['F' + str(i)] = 'CL'
        if isnull(wc['D' + str(i)].value):  # напрям вітру
            if i != 2:  # у порожніх полях вказується напрям за попереднє число місяця
                wc['D' + str(i)] = wc['D' + str(i - 1)].value
            else:  # першому елементу таблиці надається напрям за найближче наступне число місяця
                for j in range(2, 1439):
                    if not isnull(wc['D' + str(j)].value):
                        wc['D' + str(i)] = wc['D' + str(j)].value
                        break
        wc['I' + str(i)] = randint(0, 100)
        wc['G' + str(i)] = columns[0][i]
        wc['K' + str(i)] = columns[1][i]

    wb.save(months_value)


# інтерполяція даних
def data_interpolation(wc, interp_method, interp_value):
    columns = get_columns(wc, data.ic_names)
    for i in range(2):
        ic = pd.Series(columns[i])
        if isnull(ic[0]):
            ic[0] = round(ic.mean())
        if interp_method == 'linear':
            ic = round(ic.interpolate())
        if interp_method == 'polynomial':
            ic = round(ic.interpolate(method=interp_method, order=interp_value))
        if interp_method == 'pad':
            ic = round(ic.interpolate(method=interp_method, limit=interp_value))
        columns[i] = ic
    return columns


# lab 2


# 2.1 визначення часового проміжку
def get_rows(rows, day, hours='h', minutes='m'):
    day_rows = []
    day_rows1 = []
    for row in rows:
        if row[1] == 'UTC':
            day_rows.append(row)
            day_rows1.append(row)
        if day == row[0]:
            if hours != 'h':
                time = hours + ':' + minutes + ":00"
                if str(row[1]) == time:
                    day_rows1.append(row)
                    return day_rows1
            day_rows.append(row)
    return day_rows


# 2.2 температурні умови регіону (діаграма)
def t_conditions(wc):
    column = []
    for wc_item in wc['C']:
        column.append(wc_item.value)
    column.pop(0)
    time = []
    for i in range(len(column)):
        # columns[1][i] = str(columns[1][i]) # wc['B']
        time.append(i)
    plt.bar(time, column)
    plt.xlabel("Time (1 - first_day/0:00 -> 1490 - last_day/23:30)")
    plt.ylabel("Т,°C")
    plt.savefig('meteoapp/static/images/t_cond_diag.png')
    plt.close()
    # return plt.bar(['a', 'b', 'c'], [0, 1, 2])


# 2.3 тривалість температурних режимів

def duration(wc_column, x_name):
    column = []
    for wc_item in wc_column:
        column.append(wc_item.value)
    column.pop(0)

    range_and_duration = {}
    for i in range(min(column), max(column) + 1):
        range_and_duration.update([[i, 0]])

    for column_item in column:
        range_and_duration[column_item] += 0.5

    rd_list = [range_and_duration.keys(), range_and_duration.values()]

    plt.bar(rd_list[0], rd_list[1])
    plt.xlabel(x_name)
    plt.ylabel("Duration, hours")
    return rd_list


def t_duration(wc):
    rd_list = duration(wc['C'], "Т,°C")
    plt.savefig('meteoapp/static/images/t_duration_diag.png')
    plt.close()
    return rd_list


# 2.4 троянда вітрів
def wind_rose(id, city, lan1):
    months_value = 'meteoapp/cities/' + city + '/' + data.months_dict[id]
    df = pd.read_excel(months_value)

    if df['dd'][0] == data.wind_tr[0][8] or df['dd'][0] == data.wind_tr[1][8]:
        df['dd'][0] = df['dd'][1]
    for i in range(len(df.dd)):
        for j in range(len(data.wind_tr[0])):
            if df['dd'][i] == data.wind_tr[lan1][j]:
                if j == 8:
                    df['dd'][i] = df['dd'][i - 1]
                else:
                    df['dd'][i] = data.wind_degrees[j]

    # print(df['dd'])

    ax = WindroseAxes.from_ax()
    ax.bar(df.dd, df.FF, normed=True, opening=0.8, edgecolor='white')
    ax.set_legend()
    plt.savefig('meteoapp/static/images/wind_rose.png')
    plt.close()


# 2.5 тривалість режимів вітрової активності
def w_duration(wc):
    rd_list = duration(wc['E'], "Wind speed (m/s)")
    plt.savefig('meteoapp/static/images/w_duration.png')
    plt.close()
    return rd_list
